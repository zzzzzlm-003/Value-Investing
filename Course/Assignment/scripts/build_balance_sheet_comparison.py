"""
把年报 p.43 Supplemental Consolidating 与 p.49 Consolidated 两张资产负债表写入同一 Excel，便于对比。
数据来源：Deere FY2025 年报截图（p.43、p.49）。
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, "Deere-Homework-Output.xlsx")

# p.43 Supplemental Consolidating Data - 主要行（单位百万美元）
# 列顺序: 项目 | Equipment 2025 | Equipment 2024 | Financial 2025 | Financial 2024 | Eliminations 2025 | Elim 2024 | Consolidated 2025 | Consolidated 2024
P43_ASSETS = [
    ("Cash and cash equivalents", 6340, 5615, 1848, 1631, None, None, 8276, 7324),
    ("Marketable securities", 217, 125, 1194, 1029, None, None, 1411, 1154),
    ("Receivables from Financial Services", 4649, 3043, None, None, -4649, None, None, None),
    ("Trade accounts and notes receivable – net", 1316, 1257, None, None, -1899, None, 5317, 5326),
    ("Financing receivables – net", 88, 78, 44487, 44331, None, None, 44575, 44309),
    ("Financing receivables securitized – net", 1, 2, 6830, 8721, None, None, 6831, 8723),
    ("Other receivables", 1809, 2193, 594, 352, None, None, 2403, 2545),
    ("Equipment on operating leases – net", None, None, 7600, 7451, None, None, 7600, 7451),
    ("Inventories", 7406, 7093, None, None, None, None, 7406, 7093),
    ("Property and equipment – net", 8047, 7546, 32, 34, None, None, 8079, 7580),
    ("Goodwill", 4188, 3959, None, None, None, None, 4188, 3959),
    ("Other intangible assets – net", 892, 999, None, None, None, None, 892, 999),
    ("Retirement benefits", 3181, 2839, 92, 82, None, None, 3273, 2921),
    ("Deferred income taxes", 2507, 2262, -223, -176, None, None, 2284, 2086),
    ("Other assets", 2218, 2194, 1243, 712, None, None, 3461, 2906),
    ("Assets held for sale", None, None, None, None, None, None, None, 2944),
    ("Total Assets", None, None, None, None, None, None, 105996, 107320),
]

P43_LIAB_EQ = [
    ("Short-term borrowings", 414, 911, 13382, 12622, None, None, 13796, 13533),
    ("Short-term securitization borrowings", 1, 2, 6596, 8429, None, None, 6596, 8431),
    ("Payables to Equipment Operations", None, None, None, None, None, None, None, None),
    ("Accounts payable and accrued expenses", 12757, 13534, 152, 9, None, None, 13909, 14543),
    ("Deferred income taxes", 347, 434, 87, 44, None, None, 434, 478),
    ("Long-term borrowings", 8756, 6603, 34788, 36626, None, None, 43544, 43229),
    ("Retirement benefits and other liabilities", 1646, 2250, 64, 104, None, None, 1710, 2354),
    ("Liabilities held for sale", None, None, None, None, None, None, None, 1827),
    ("Total liabilities", 23291, 23734, 56069, 57838, None, None, 79989, 84395),
    ("Redeemable noncontrolling interest", 51, 82, None, None, None, None, 51, 82),
    ("Total Deere & Company stockholders' equity", 25950, 22836, None, None, None, None, 25950, 22836),
    ("Noncontrolling interests", 6, 7, None, None, None, None, 6, 7),
    ("Financial Services' equity", -7069, -7454, None, None, 7069, 7454, None, None),
    ("Adjusted total stockholders' equity", 18938, 15389, None, None, 7069, 7454, 25956, 22843),
    ("Total Liabilities and Stockholders' Equity", None, None, None, None, None, None, 105996, 107320),
]

# p.49 Consolidated Balance Sheet（与 p.43 的 Consolidated 列一致，仅展示格式不同）
P49_ASSETS = [
    ("Cash and cash equivalents", 8276, 7324),
    ("Marketable securities", 1411, 1154),
    ("Trade accounts and notes receivable – net", 5317, 5326),
    ("Financing receivables – net", 44575, 44309),
    ("Financing receivables securitized – net", 6831, 8723),
    ("Other receivables", 2403, 2545),
    ("Equipment on operating leases – net", 7600, 7451),
    ("Inventories", 7406, 7093),
    ("Property and equipment – net", 8079, 7580),
    ("Goodwill", 4188, 3959),
    ("Other intangible assets – net", 892, 999),
    ("Retirement benefits", 3273, 2921),
    ("Deferred income taxes", 2284, 2086),
    ("Other assets", 3461, 2906),
    ("Assets held for sale", None, 2944),
    ("Total Assets", 105996, 107320),
]

P49_LIAB_EQ = [
    ("Short-term borrowings", 13796, 13533),
    ("Short-term securitization borrowings", 6596, 8431),
    ("Accounts payable and accrued expenses", 13909, 14543),
    ("Deferred income taxes", 434, 478),
    ("Long-term borrowings", 43544, 43229),
    ("Retirement benefits and other liabilities", 1710, 2354),
    ("Liabilities held for sale", None, 1827),
    ("Total liabilities", 79989, 84395),
    ("Redeemable noncontrolling interest", 51, 82),
    ("Total Deere & Company stockholders' equity", 25950, 22836),
    ("Noncontrolling interests", 6, 7),
    ("Total stockholders' equity", 25956, 22843),
    ("Total Liabilities and Stockholders' Equity", 105996, 107320),
]


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)) and v == int(v):
        return int(v)
    return v


def write_p43(ws):
    ws.append(["SUPPLEMENTAL CONSOLIDATING DATA — CONDENSED BALANCE SHEETS (p.43)", ""])
    ws.append(["($ in millions)", ""])
    ws.append([])
    h = ["Line Item", "Equip 2025", "Equip 2024", "Financial 2025", "Financial 2024", "Elim 2025", "Elim 2024", "Consol 2025", "Consol 2024"]
    ws.append(h)
    for row in P43_ASSETS:
        ws.append([row[0]] + [_fmt(x) for x in row[1:]])
    ws.append([])
    ws.append(["LIABILITIES AND STOCKHOLDERS' EQUITY"])
    ws.append(h)
    for row in P43_LIAB_EQ:
        ws.append([row[0]] + [_fmt(x) for x in row[1:]])
    ws.append([])
    ws.append(["Note: Financial Services' equity is (7,069) in Equipment Operations; +7,069 in Eliminations (footnote 10)."])


def write_p49(ws):
    ws.append(["CONSOLIDATED BALANCE SHEETS (p.49)", ""])
    ws.append(["As of November 2, 2025 and October 27, 2024 — $ in millions", ""])
    ws.append([])
    ws.append(["ASSETS", "2025", "2024"])
    for row in P49_ASSETS:
        ws.append([row[0], _fmt(row[1]), _fmt(row[2])])
    ws.append([])
    ws.append(["LIABILITIES AND STOCKHOLDERS' EQUITY", "2025", "2024"])
    for row in P49_LIAB_EQ:
        ws.append([row[0], _fmt(row[1]), _fmt(row[2])])


def write_comparison(ws):
    ws.append(["Comparison: p.43 Consolidated column vs p.49", ""])
    ws.append(["Item", "p.43 Consolidated 2025", "p.49 2025", "Match?"])
    ws.append(["Total Assets", 105996, 105996, "Yes"])
    ws.append(["Total Liabilities", 79989, 79989, "Yes"])
    ws.append(["Total Stockholders' Equity", 25956, 25956, "Yes"])
    ws.append([])
    ws.append(["Key: p.43 shows Equipment | Financial | Eliminations | Consolidated. p.49 is the same as Consolidated."])
    ws.append(["Equipment Operations equity is reduced by 'Financial Services equity' (7,069); Eliminations add it back for consolidation."])


def main():
    if os.path.exists(OUT):
        wb = openpyxl.load_workbook(OUT)
    else:
        wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    for s in ["BS_p43_Supplemental_Consolidati", "BS_p49_Consolidated", "BS_Comparison"]:
        if s in wb.sheetnames:
            del wb[s]

    ws43 = wb.create_sheet("BS_p43_Supplemental_Consolidati")
    ws49 = wb.create_sheet("BS_p49_Consolidated")
    ws_cmp = wb.create_sheet("BS_Comparison")
    write_p43(ws43)
    write_p49(ws49)
    write_comparison(ws_cmp)
    for ws in [ws43, ws49, ws_cmp]:
        for c in range(1, min(ws.max_column + 1, 12)):
            ws.cell(1, c).font = Font(bold=True)
    wb.save(OUT)
    print("已写入:", OUT)


if __name__ == "__main__":
    main()
