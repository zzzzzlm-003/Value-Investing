"""
Question 2: Farm equipment competitors revenue shares 2006-2024.
- Load Industry sheet from Homework-Data-2026.xlsx
- Convert Kubota (JPY) and Claas (EUR) to USD using yearly average exchange rates
- Plot stacked bar / line chart of revenue shares
- Save chart: (1) image file, (2) embed in Excel Industry sheet
"""
import os
import sys

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

# Yearly average exchange rates (sources: FRED, Fed H.10, ECB)
# JPY per 1 USD - to convert Kubota JPY to USD: divide by this
JPY_USD = {
    2006: 116.30, 2007: 117.75, 2008: 103.36, 2009: 93.57, 2010: 87.78,
    2011: 79.81, 2012: 79.79, 2013: 97.60, 2014: 105.94, 2015: 121.04,
    2016: 108.78, 2017: 112.17, 2018: 110.42, 2019: 109.01, 2020: 106.78,
    2021: 109.75, 2022: 131.50, 2023: 140.50, 2024: 149.50,
}
# USD per 1 EUR - to convert Claas EUR to USD: multiply by this
EUR_USD = {
    2006: 1.256, 2007: 1.371, 2008: 1.471, 2009: 1.395, 2010: 1.326,
    2011: 1.392, 2012: 1.285, 2013: 1.328, 2014: 1.329, 2015: 1.109,
    2016: 1.107, 2017: 1.130, 2018: 1.181, 2019: 1.120, 2020: 1.142,
    2021: 1.183, 2022: 1.053, 2023: 1.082, 2024: 1.087,
}

# Deere missing years - from Deere Annual Report (Equipment Net Sales, $mn)
# 2006 extrapolated; 2021-2024 from Supplemental/Selected Financial Data
DEERE_FILL = {2006: 15200, 2021: 39737, 2022: 47917, 2023: 55565, 2024: 44759}

# Farm revenue as % of total (for pure-farm share calculation)
# Deere: PPA+SAT / equipment ~71%; CNH: ag segment ~77%; AGCO: ~95%; Kubota: ~79%; Claas: ~98%
FARM_PCT = {"Deere": 0.71, "Kubota": 0.79, "CNH": 0.77, "AGCO": 0.95, "Claas": 0.98}


def load_industry(base_path):
    df = pd.read_excel(os.path.join(base_path, "Homework-Data-2026.xlsx"), sheet_name="Industry", header=None)
    # Parse: rows 4-22 = data. Col 2 = year (2006, 2016-2024); 3=Deere, 4=Kubota, 5=CNH, 6=AGCO, 7=Claas
    # Years: row4=2006, row5-13=2007-2015, row14=2016, row15=2017, ..., row22=2024
    years = [2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024]
    data = []
    for i, yr in enumerate(years):
        row = df.iloc[4 + i]
        def num(v):
            if pd.isna(v): return None
            try: return float(v)
            except: return None
        data.append({
            "Year": yr,
            "Deere": num(row[3]),
            "Kubota_JPY": num(row[4]),
            "CNH": num(row[5]),
            "AGCO": num(row[6]),
            "Claas_EUR": num(row[7]),
        })
    return pd.DataFrame(data)


def convert_to_usd(df):
    df = df.copy()
    df["Deere_USD"] = df.apply(lambda r: DEERE_FILL.get(r["Year"], r["Deere"]) if pd.isna(r["Deere"]) else r["Deere"], axis=1)
    df["Kubota_USD"] = df["Kubota_JPY"] / df["Year"].map(JPY_USD)
    df["CNH_USD"] = df["CNH"]
    df["AGCO_USD"] = df["AGCO"]
    df["Claas_USD"] = df["Claas_EUR"] * df["Year"].map(EUR_USD)
    
    cols = ["Deere_USD", "Kubota_USD", "CNH_USD", "AGCO_USD", "Claas_USD"]
    df["Total_USD"] = df[cols].sum(axis=1)
    for c in cols:
        df[c.replace("_USD", "_share")] = (df[c] / df["Total_USD"] * 100).round(1)
    return df


def plot_shares(df, out_path, title_suffix="", share_cols=None):
    years = df["Year"].astype(int)
    companies = ["Deere", "Kubota", "CNH", "AGCO", "Claas"]
    if share_cols is None:
        share_cols = [f"{c}_share" for c in companies]
    colors = ["#367c2b", "#e35d2b", "#1a5276", "#f39c12", "#8e44ad"]
    
    fig, ax = plt.subplots(figsize=(12, 6))
    bottom = None
    for i, (comp, col) in enumerate(zip(companies, share_cols)):
        vals = df[col].values
        if bottom is None:
            ax.bar(years, vals, label=comp, color=colors[i])
            bottom = vals
        else:
            ax.bar(years, vals, bottom=bottom, label=comp, color=colors[i])
            bottom = bottom + vals
    
    ax.set_xlabel("Year")
    ax.set_ylabel("Revenue Share (%)")
    ax.set_title(f"Farm Equipment Industry: Revenue Shares 2006-2024{title_suffix}")
    ax.legend(loc="upper right")
    ax.set_xticks(years)
    ax.set_xticklabels(years, rotation=45)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"Chart saved: {out_path}")


def compute_farm_only(df):
    """Farm-only revenue and shares: apply farm % to each company's total revenue."""
    df = df.copy()
    for c in ["Deere", "Kubota", "CNH", "AGCO", "Claas"]:
        usd_col = f"{c}_USD"
        df[f"{c}_Farm_USD"] = df[usd_col] * FARM_PCT[c]
    farm_cols = [f"{c}_Farm_USD" for c in ["Deere", "Kubota", "CNH", "AGCO", "Claas"]]
    df["Total_Farm_USD"] = df[farm_cols].sum(axis=1)
    for c in ["Deere", "Kubota", "CNH", "AGCO", "Claas"]:
        df[f"{c}_Farm_share"] = (df[f"{c}_Farm_USD"] / df["Total_Farm_USD"] * 100).round(1)
    return df


IMG_DIR = "images"
OUTPUT_XLSX = "Deere-Homework-Output.xlsx"


def save_to_output_excel(df, base_path):
    """Save share tables to consolidated output Excel."""
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    companies = ["Deere", "Kubota", "CNH", "AGCO", "Claas"]
    cols_total = ["Year"] + [f"{c}_share" for c in companies]
    cols_farm = ["Year"] + [f"{c}_Farm_share" for c in companies]
    out_path = os.path.join(base_path, OUTPUT_XLSX)
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        for s in ["Revenue_Shares_Total", "Revenue_Shares_Farm"]:
            if s in wb.sheetnames:
                del wb[s]
    else:
        wb = __import__("openpyxl").Workbook()
        wb.remove(wb.active)
    ws1 = wb.create_sheet("Revenue_Shares_Total")
    ws2 = wb.create_sheet("Revenue_Shares_Farm")
    for r in dataframe_to_rows(df[cols_total], index=False, header=True):
        ws1.append(r)
    for r in dataframe_to_rows(df[cols_farm], index=False, header=True):
        ws2.append(r)
    wb.save(out_path)
    print(f"Saved to {OUTPUT_XLSX}")


def main():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    img_dir = os.path.join(base, IMG_DIR)
    os.makedirs(img_dir, exist_ok=True)
    df = load_industry(base)
    df = convert_to_usd(df)
    df = compute_farm_only(df)
    
    # Chart 1: Total revenue shares
    img_path = os.path.join(img_dir, "industry_revenue_shares.png")
    plot_shares(df, img_path, title_suffix="\n(Total company revenue; currency converted using yearly avg rates)")
    
    # Chart 2: Pure farm revenue shares
    farm_img_path = os.path.join(img_dir, "industry_revenue_shares_farm_only.png")
    farm_cols = [f"{c}_Farm_share" for c in ["Deere", "Kubota", "CNH", "AGCO", "Claas"]]
    plot_shares(df, farm_img_path, title_suffix="\n(Pure farm business only; farm % applied to each company)", share_cols=farm_cols)
    
    save_to_output_excel(df, base)
    
    # Print both tables
    print("\n--- Revenue Shares: Total Company (%) ---")
    print(df[["Year", "Deere_share", "Kubota_share", "CNH_share", "AGCO_share", "Claas_share"]].to_string(index=False))
    print("\n--- Revenue Shares: Pure Farm Business (%) ---")
    print(df[["Year", "Deere_Farm_share", "Kubota_Farm_share", "CNH_Farm_share", "AGCO_Farm_share", "Claas_Farm_share"]].to_string(index=False))
    
    return df


if __name__ == "__main__":
    main()
