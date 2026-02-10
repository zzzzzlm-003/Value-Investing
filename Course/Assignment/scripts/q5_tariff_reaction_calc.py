"""
Q5: 股价对关税宣布的反应 - 计算过程展示
数据来源: Homework-Data-2026.xlsx, sheet "Deere in 2024 and 2025"
特朗普2025年多次重大关税宣布，分别计算每次后的股价反应
"""
import os
import pandas as pd
from datetime import timedelta

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE, "Homework-Data-2026.xlsx")
IMG_DIR = "images"
OUTPUT_XLSX = "Deere-Homework-Output.xlsx"
SHEET = "Deere in 2024 and 2025"
DAYS_AFTER = 14  # 约两周

# 特朗普2025年重大关税宣布/生效日期（公开报道）
TARIFF_EVENTS = [
    ("2025-01-20", "就职日 America First 贸易政策备忘录"),
    ("2025-02-01", "对加拿大/墨西哥/中国加征关税"),
    ("2025-02-10", "钢铁铝材关税提至25%"),
    ("2025-03-26", "进口汽车/轻卡25%关税"),
    ("2025-04-02", "全球10%基准关税(白宫宣布)"),
    ("2025-04-09", "中国关税提至125%+ 其他暂停90天"),
    ("2025-08-07", "69国互惠关税生效(10-41%)"),
]


def analyze_one_event(df, announce_dt, event_desc, days_after=14):
    """对单次关税事件：两周内最低价、此后最高价、收益率"""
    window_end = announce_dt + timedelta(days=days_after)
    after = df[df["Date"] >= announce_dt]
    if len(after) < 3:
        return None
    early = after[after["Date"] <= window_end]
    late = after[after["Date"] > window_end]
    if len(early) == 0:
        return None
    p_min = early["Stock Price"].min()
    date_min = early.loc[early["Stock Price"].idxmin(), "Date"].date()
    if len(late) == 0:
        p_max = early["Stock Price"].max()
        date_max = early.loc[early["Stock Price"].idxmax(), "Date"].date()
        ret = 0.0
    else:
        p_max = late["Stock Price"].max()
        date_max = late.loc[late["Stock Price"].idxmax(), "Date"].date()
        ret = (p_max / p_min) - 1 if p_min > 0 else 0
    return {
        "event": event_desc,
        "announce": announce_dt.date(),
        "window_end": window_end.date(),
        "p_min": p_min,
        "date_min": date_min,
        "p_max": p_max,
        "date_max": date_max,
        "return_pct": ret * 100,
    }


def main():
    print("=" * 70)
    print("Q5 计算过程：股价对特朗普多次关税宣布的反应")
    print("=" * 70)

    # 1. 读取数据
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET)
    df = df.rename(columns={df.columns[1]: "Date"})
    df["Date"] = pd.to_datetime(df["Date"])
    df = df[["Date", "Stock Price", "Revenues ($mn)"]].dropna(subset=["Stock Price"])
    df = df.sort_values("Date").reset_index(drop=True)

    print("\n【数据】")
    print(f"  文件: Homework-Data-2026.xlsx, Sheet: {SHEET}")
    print(f"  日度股价: {len(df)} 行, 日期范围 {df['Date'].min().date()} ~ {df['Date'].max().date()}")
    print(f"  方法: 每次关税宣布后，取「约两周内」最低价、此后最高价，收益率=(最高/最低)-1")

    # 2. 逐次关税事件计算
    results = []
    for date_str, desc in TARIFF_EVENTS:
        announce = pd.Timestamp(date_str)
        r = analyze_one_event(df, announce, desc, DAYS_AFTER)
        if r:
            results.append(r)

    # 3. 打印各次结果
    print("\n【各次关税宣布后的 Deere 股价反应】")
    print("-" * 70)
    for r in results:
        print(f"\n  {r['announce']} | {r['event']}")
        print(f"    窗口: 宣布日 ~ {r['window_end']} (两周)")
        print(f"    两周内最低: ${r['p_min']:.2f} ({r['date_min']})")
        print(f"    此后最高:   ${r['p_max']:.2f} ({r['date_max']})")
        print(f"    收益率:    {r['return_pct']:.2f}%  [= ({r['p_max']:.2f}/{r['p_min']:.2f})-1]")

    # 4. 汇总表
    print("\n" + "=" * 70)
    print("汇总表")
    print("=" * 70)
    tbl = pd.DataFrame(results)[["announce", "event", "p_min", "date_min", "p_max", "date_max", "return_pct"]]
    tbl.columns = ["宣布日", "事件", "最低价", "最低日", "最高价", "最高日", "收益率(%)"]
    print(tbl.to_string(index=False))

    # 5. 季度收入
    rev = df.drop_duplicates(subset=["Revenues ($mn)"], keep="first")[["Date", "Revenues ($mn)"]].sort_values("Date")
    print("\n【季度收入】财年 Q1=Nov-Jan, Q2=Feb-Apr, Q3=May-Jul, Q4=Aug-Oct")
    for _, r in rev.iterrows():
        print(f"  {r['Date'].strftime('%Y-%m-%d')}: ${r['Revenues ($mn)']:,.0f} mn")

    # 6. 写入 consolidated Excel
    os.makedirs(os.path.join(BASE, IMG_DIR), exist_ok=True)
    out_path = os.path.join(BASE, OUTPUT_XLSX)
    excel_help = pd.DataFrame([
        ["步骤", "说明"],
        ["1", "数据: Homework-Data-2026.xlsx, sheet 'Deere in 2024 and 2025'"],
        ["2", "对每次关税宣布: 两周内=MIN(股价), 两周后=MAX(股价)"],
        ["3", "收益率 = (最高价/最低价) - 1"],
        ["4 最低价", "=MINIFS(C:C,B:B,\">=宣布日\",B:B,\"<=宣布日+14\")"],
        ["5 最高价", "=MAXIFS(C:C,B:B,\">宣布日+14\")"],
    ], columns=["步骤", "说明"])
    from openpyxl import load_workbook
    q5_sheets = ["Q5_Tariff_Reaction", "Q5_Quarterly_Rev", "Q5_Formula"]
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        for s in q5_sheets:
            if s in wb.sheetnames:
                del wb[s]
        wb.save(out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="a" if os.path.exists(out_path) else "w") as w:
        tbl.to_excel(w, sheet_name="Q5_Tariff_Reaction", index=False)
        rev.to_excel(w, sheet_name="Q5_Quarterly_Rev", index=False)
        excel_help.to_excel(w, sheet_name="Q5_Formula", index=False)
    print(f"\n结果已写入: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
