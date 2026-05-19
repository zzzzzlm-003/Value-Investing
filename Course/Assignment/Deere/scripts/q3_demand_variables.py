"""
Question 3: Farm equipment demand determinants.
- Fetch from FRED only (no built-in estimates)
- Store in Excel, Plot: YoY (lines) + Deere sales (bars)
Data sources: FRED (Federal Reserve Economic Data), stlouisfed.org
"""
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG_DIR = "images"
OUTPUT_XLSX = "Deere-Homework-Output.xlsx"

# FRED series IDs - all from FRED, verified
# Source: https://fred.stlouisfed.org
FRED_SERIES = {
    "Corn_USD_MT": "PMAIZMTUSDA",           # Global price of Corn, USD/metric ton (IMF)
    "Soybean_USD_MT": "PSOYBUSDA",          # Global price of Soybeans, USD/metric ton (IMF)
    "Oil_WTI_USD_bbl": "MCOILWTICO",        # WTI Crude Oil, USD/bbl (EIA, monthly)
    "FarmRealEstate_mn_USD": "BOGZ1FL135035005A",  # Farm Business Real Estate USDA, Level $mn (Fed Z.1)
    "Labor_FarmWeeklyEarn_USD": "LEU0254557500A",  # Median weekly earnings, Farming/fishing/forestry (BLS)
    "NatGas_USD_per_MMBtu": "DHHNGSP",      # Henry Hub Natural Gas, $/MMBtu (EIA, daily)
    "Energy_Price_Index": "PNRGINDEXM",     # Global Price of Energy Index, 2016=100 (IMF, monthly)
}

# Deere equipment net sales ($mn) - from Industry sheet + Financials
DEERE_SALES = {
    2006: 15200, 2007: 16454, 2008: 20985, 2009: 18122, 2010: 19868,
    2011: 24094, 2012: 27123, 2013: 29132, 2014: 26380, 2015: 19812,
    2016: 18847, 2017: 20167, 2018: 23191, 2019: 23666, 2020: 22325,
    2021: 39737, 2022: 47917, 2023: 55565, 2024: 44759,
}


def fetch_fred(series_dict):
    """Fetch from FRED REST API. Requires FRED_API_KEY."""
    import urllib.request
    import json
    key = os.environ.get("FRED_API_KEY", "")
    if not key:
        return None
    out = {}
    for name, sid in series_dict.items():
        try:
            url = f"https://api.stlouisfed.org/fred/series/observations?series_id={sid}&api_key={key}&file_type=json&observation_start=2000-01-01"
            with urllib.request.urlopen(url, timeout=15) as resp:
                data = json.load(resp)
            obs = data.get("observations", [])
            if not obs:
                continue
            dates = []
            vals = []
            for o in obs:
                v = o.get("value")
                if v and v != ".":
                    dates.append(pd.to_datetime(o["date"]))
                    vals.append(float(v))
            if vals:
                s = pd.Series(vals, index=pd.DatetimeIndex(dates))
                out[name] = s
        except Exception as e:
            print(f"  Skip {sid}: {e}")
    return out if out else None


def fetch_and_build():
    """Fetch all data from FRED. No built-in fallback."""
    raw = fetch_fred(FRED_SERIES)
    if raw is None or len(raw) < 2:
        raise SystemExit("FRED_API_KEY required. Set it and re-run.")
    print("Fetching from FRED...")
    dfs = []
    for name, s in raw.items():
        try:
            ann = s.resample("YE").mean().dropna()
        except Exception:
            ann = s.resample("Y").mean().dropna()
        yrs = ann.index.year if hasattr(ann.index, 'year') else [pd.Timestamp(x).year for x in ann.index]
        dfs.append(pd.DataFrame({"Year": yrs, name: ann.values}))
    df = dfs[0]
    for d in dfs[1:]:
        df = df.merge(d, on="Year", how="outer")
    df = df.sort_values("Year").reset_index(drop=True)
    df = df[df["Year"].between(2000, 2030)]
    return df


def compute_yoy(df):
    for c in df.columns:
        if c == "Year" or c == "Deere_Sales_mn": continue
        if df[c].dtype in [np.float64, np.int64]:
            df[f"{c}_YoY"] = df[c].pct_change(fill_method=None) * 100
    return df


def _save_to_output_excel(df, base):
    path = os.path.join(base, OUTPUT_XLSX)
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    if os.path.exists(path):
        wb = load_workbook(path)
        if "Demand_Variables" in wb.sheetnames:
            del wb["Demand_Variables"]
    else:
        wb = __import__("openpyxl").Workbook()
        wb.remove(wb.active)
    ws = wb.create_sheet("Demand_Variables")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(path)
    print(f"Data saved to {OUTPUT_XLSX}")


def plot_combined(df, out_path):
    df = df[df["Year"].between(2006, 2024)].copy()
    years = df["Year"].astype(int)
    
    fig, ax1 = plt.subplots(figsize=(14, 8))
    ax2 = ax1.twinx()
    
    # Deere sales as bars (left axis)
    sales = [DEERE_SALES.get(y, np.nan) for y in years]
    bars = ax1.bar(years - 0.3, sales, width=0.5, color="#367c2b", alpha=0.7, label="Deere Equipment Sales ($mn)")
    ax1.set_ylabel("Deere Equipment Net Sales ($ millions)", color="#367c2b")
    ax1.tick_params(axis="y", labelcolor="#367c2b")
    ax1.set_ylim(0, max(sales) * 1.15)
    
    # YoY lines (right axis)
    yoy_cols = [c for c in df.columns if c.endswith("_YoY")]
    colors = ["#e74c3c", "#3498db", "#9b59b6", "#e67e22", "#1abc9c", "#2ecc71"]
    for i, col in enumerate(yoy_cols[:6]):
        label = col.replace("_YoY", "").replace("_", " ")
        ax2.plot(years, df[col], "o-", color=colors[i % len(colors)], label=label, linewidth=1.5, markersize=4)
    ax2.axhline(0, color="gray", linestyle="--", alpha=0.5)
    ax2.set_ylabel("YoY Change (%)", color="#333")
    ax2.tick_params(axis="y", labelcolor="#333")
    
    ax1.set_xlabel("Year")
    ax1.set_title("Farm Equipment Demand: Deere Sales vs. Key Variables (YoY %)\nSource: FRED (Corn, Soybean, Farm Real Estate, Farm Labor Earnings, Oil, NatGas, Energy Index)")
    ax1.legend(loc="upper left")
    ax2.legend(loc="upper right", fontsize=8)
    ax1.set_xticks(years)
    ax1.set_xticklabels(years, rotation=45)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"Chart saved: {out_path}")


def main():
    os.makedirs(os.path.join(BASE, IMG_DIR), exist_ok=True)
    df = fetch_and_build()
    if "Year" not in df.columns and df.index.name is None:
        df["Year"] = range(2006, 2006 + len(df))
    df = compute_yoy(df)
    df["Deere_Sales_mn"] = df["Year"].map(DEERE_SALES)
    
    _save_to_output_excel(df, BASE)
    plot_combined(df, os.path.join(BASE, IMG_DIR, "demand_deere_combined.png"))
    return df


if __name__ == "__main__":
    main()
